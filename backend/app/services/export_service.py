from __future__ import annotations

import csv
import io
from typing import Iterable

from openpyxl import Workbook


def build_csv_bytes(rows: Iterable[dict]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "student_name", "student_id", "project_title", "practicality", "innovation", "total", "needs_review"])
    for row in rows:
        writer.writerow(
            [
                row.get("id"),
                row.get("student_name"),
                row.get("student_id"),
                row.get("project_title"),
                row.get("practicality_score"),
                row.get("innovation_score"),
                row.get("total_score"),
                1 if row.get("needs_human_review") else 0,
            ]
        )
    return output.getvalue().encode("utf-8-sig")


def build_xlsx_bytes(rows: Iterable[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "scores"
    ws.append(["id", "student_name", "student_id", "project_title", "practicality", "innovation", "total", "needs_review"])
    for row in rows:
        ws.append(
            [
                row.get("id"),
                row.get("student_name"),
                row.get("student_id"),
                row.get("project_title"),
                row.get("practicality_score"),
                row.get("innovation_score"),
                row.get("total_score"),
                1 if row.get("needs_human_review") else 0,
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_teacher_score_csv_bytes(rows: Iterable[dict]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "submission_id",
            "assignment_id",
            "course_name",
            "assignment_title",
            "student_id",
            "student_name",
            "group_name",
            "project_name",
            "teacher_student_id",
            "teacher_role",
            "assigned",
            "scored",
            "innovation",
            "completeness",
            "code_quality",
            "demo",
            "contribution",
            "total",
            "comment",
            "updated_at",
        ]
    )
    for row in rows:
        writer.writerow(
            [
                row.get("submission_id"),
                row.get("assignment_id"),
                row.get("course_name"),
                row.get("assignment_title"),
                row.get("student_id"),
                row.get("student_name"),
                row.get("group_name"),
                row.get("project_name"),
                row.get("teacher_student_id"),
                row.get("teacher_role"),
                1 if row.get("assigned") else 0,
                1 if row.get("scored") else 0,
                row.get("innovation_score"),
                row.get("completeness_score"),
                row.get("code_quality_score"),
                row.get("demo_score"),
                row.get("contribution_score"),
                row.get("total_score"),
                row.get("comment"),
                row.get("updated_at"),
            ]
        )
    return output.getvalue().encode("utf-8-sig")


def build_teacher_score_xlsx_bytes(summary_rows: Iterable[dict], score_rows: Iterable[dict]) -> bytes:
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "summary"
    ws_summary.append(
        [
            "submission_id",
            "assignment_id",
            "course_name",
            "assignment_title",
            "student_id",
            "student_name",
            "group_name",
            "project_name",
            "status",
            "completeness_status",
            "assigned_teacher_count",
            "score_count",
            "average_total_score",
            "average_innovation_score",
            "average_completeness_score",
            "average_code_quality_score",
            "average_demo_score",
            "average_contribution_score",
            "updated_at",
        ]
    )
    for row in summary_rows:
        ws_summary.append(
            [
                row.get("submission_id"),
                row.get("assignment_id"),
                row.get("course_name"),
                row.get("assignment_title"),
                row.get("student_id"),
                row.get("student_name"),
                row.get("group_name"),
                row.get("project_name"),
                row.get("status"),
                row.get("completeness_status"),
                row.get("assigned_teacher_count"),
                row.get("score_count"),
                row.get("average_total_score"),
                row.get("average_innovation_score"),
                row.get("average_completeness_score"),
                row.get("average_code_quality_score"),
                row.get("average_demo_score"),
                row.get("average_contribution_score"),
                row.get("updated_at"),
            ]
        )

    ws_scores = wb.create_sheet("teacher_scores")
    ws_scores.append(
        [
            "submission_id",
            "assignment_id",
            "course_name",
            "assignment_title",
            "student_id",
            "student_name",
            "group_name",
            "project_name",
            "teacher_student_id",
            "teacher_role",
            "assigned",
            "scored",
            "innovation",
            "completeness",
            "code_quality",
            "demo",
            "contribution",
            "total",
            "comment",
            "updated_at",
        ]
    )
    for row in score_rows:
        ws_scores.append(
            [
                row.get("submission_id"),
                row.get("assignment_id"),
                row.get("course_name"),
                row.get("assignment_title"),
                row.get("student_id"),
                row.get("student_name"),
                row.get("group_name"),
                row.get("project_name"),
                row.get("teacher_student_id"),
                row.get("teacher_role"),
                1 if row.get("assigned") else 0,
                1 if row.get("scored") else 0,
                row.get("innovation_score"),
                row.get("completeness_score"),
                row.get("code_quality_score"),
                row.get("demo_score"),
                row.get("contribution_score"),
                row.get("total_score"),
                row.get("comment"),
                row.get("updated_at"),
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
